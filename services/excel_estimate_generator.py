#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel генератор смет для EventBot AI v2.0"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional

class ExcelEstimateGenerator:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)
        
        # Проверяем наличие openpyxl
        try:
            import openpyxl
            self.excel_available = True
            self.logger.info("Excel генератор готов")
        except ImportError:
            self.excel_available = False
            self.logger.warning("openpyxl не установлен, Excel файлы не будут создаваться")
    
    def create_estimate(self, estimate_data: Dict, request_data: Dict) -> Optional[str]:
        """Создание Excel файла сметы"""
        
        if not self.excel_available:
            self.logger.warning("Excel генератор недоступен")
            return None
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Создаем новую книгу
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Смета"
            
            # Стили
            header_font = Font(name='Arial', size=14, bold=True)
            title_font = Font(name='Arial', size=16, bold=True)
            normal_font = Font(name='Arial', size=11)
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
            
            # Заголовок
            ws.merge_cells('A1:F1')
            ws['A1'] = "СМЕТА НА КЕЙТЕРИНГОВОЕ ОБСЛУЖИВАНИЕ"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Информация о компании
            ws['A3'] = "Ресторан РестДеливери"
            ws['A3'].font = header_font
            ws['A4'] = "Профессиональный банкетное обслуживание"
            ws['A5'] = "Тел: +7 (XXX) XXX-XX-XX"
            ws['A6'] = "Email: info@restdelivery.ru"
            
            # Информация о мероприятии
            row = 8
            ws[f'A{row}'] = "ИНФОРМАЦИЯ О МЕРОПРИЯТИИ"
            ws[f'A{row}'].font = header_font
            row += 1
            
            event_info = [
                ("Тип мероприятия:", request_data.get('event_type', 'Не указан').title()),
                ("Количество гостей:", str(request_data.get('guest_count', 'Не указано'))),
                ("Длительность:", f"{request_data.get('duration', 3)} часа"),
                ("Дата создания сметы:", datetime.now().strftime('%d.%m.%Y %H:%M'))
            ]
            
            for label, value in event_info:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
            
            # Смета
            row += 2
            ws[f'A{row}'] = "ДЕТАЛЬНАЯ СМЕТА"
            ws[f'A{row}'].font = header_font
            row += 1
            
            # Заголовки таблицы
            headers = ['№', 'Наименование', 'Единица', 'Количество', 'Цена', 'Сумма']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            
            # Позиции меню из estimate_data
            menu_items = estimate_data.get('menu_items', [])
            total_menu = 0
            for i, item in enumerate(menu_items, 1):
                name = item.get('name', '')
                unit = item.get('unit', 'шт')
                qty = int(round(item.get('quantity', 0)))
                price = int(round(item.get('price', 0)))
                amount = int(round(item.get('total_cost', qty * price)))
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=name)
                ws.cell(row=row, column=3, value=unit)
                ws.cell(row=row, column=4, value=qty)
                ws.cell(row=row, column=5, value=price)
                ws.cell(row=row, column=6, value=amount)
                # Применяем стили
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.font = normal_font
                    if col >= 4:  # Числовые колонки
                        cell.alignment = Alignment(horizontal='right')
                total_menu += amount
                row += 1
            # Итого по меню
            ws.cell(row=row, column=5, value="ИТОГО МЕНЮ:")
            ws.cell(row=row, column=6, value=int(round(total_menu)))
            ws.cell(row=row, column=5).font = header_font
            ws.cell(row=row, column=6).font = header_font
            row += 2
            # Услуги
            service_cost = int(round(estimate_data.get('service_cost', total_menu * 0.2)))
            total_cost = int(round(estimate_data.get('total_cost', total_menu + service_cost)))
            services = [
                ("Обслуживающий персонал", int(round(service_cost * 0.7))),
                ("Доставка и логистика", int(round(service_cost * 0.2))),
                ("Оборудование", int(round(service_cost * 0.1)))
            ]
            ws[f'A{row}'] = "УСЛУГИ"
            ws[f'A{row}'].font = header_font
            row += 1
            for service, cost in services:
                ws.cell(row=row, column=2, value=service)
                ws.cell(row=row, column=6, value=int(round(cost)))
                row += 1
            # Общий итог
            row += 1
            ws.cell(row=row, column=5, value="ОБЩАЯ СУММА:")
            ws.cell(row=row, column=6, value=int(round(total_cost)))
            ws.cell(row=row, column=5).font = title_font
            ws.cell(row=row, column=6).font = title_font
            row += 1
            cost_per_guest = int(round(total_cost / max(1, request_data.get('guest_count', 1))))
            ws.cell(row=row, column=5, value="Стоимость на человека:")
            ws.cell(row=row, column=6, value=f"{cost_per_guest} руб")
            # Настройка ширины колонок
            column_widths = [5, 30, 12, 12, 12, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            # Сохранение файла
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"smeta_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            wb.save(filepath)
            
            self.logger.info(f"Excel файл создан: {filename}")
            return str(filepath)
        except Exception as e:
            self.logger.error(f"Ошибка создания Excel файла: {e}")
            return None
